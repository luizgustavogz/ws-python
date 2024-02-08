# openpyxl - ler e alterar dados de uma planilha
# Com essa biblioteca será possível ler e escrever dados em células específicas,
# formatar células, inserir gráficos, criar fórmulas, adicionar imagens e outros
# elementos gráficos às suas planilhas. Ela é útil para automatizar tarefas
# envolvendo planilhas do Excel, como a criação de relatórios e análise de
# dados e/ou facilitando a manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando o arquivo
workbook = load_workbook(WORKBOOK_PATH)

# Nomeando a planilha
sheet_name = 'Minha Planilha'

# Selecionando a planilha
worksheet: Worksheet = workbook[sheet_name]


for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, '22')
    print()

# worksheet['B3'].value = 14

workbook.save(WORKBOOK_PATH)
